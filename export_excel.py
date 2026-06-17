"""
export_excel.py
---------------
Reads ``attendance_log_temp.csv`` and exports a formatted Excel file with:
  • Auto-width columns
  • Auto-height rows
  • Embedded snapshot thumbnails in the Photo column

Run:
    python export_excel.py
"""

import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXPORTS_DIR   = "exports"          # All CSVs and Excel exports live here
LOG_FILE_PATH = os.path.join(EXPORTS_DIR, "attendance_log_temp.csv")
OUTPUT_DIR    = EXPORTS_DIR
THUMB_W       = 80                # Thumbnail width in pixels (for cell embedding)
THUMB_H       = 80
ROW_HEIGHT_PX = 65                # Excel row height (points) for photo rows
HEADER_COLOR  = "1F4E79"          # Dark blue header background
ALT_ROW_COLOR = "D6E4F0"         # Light blue alternate row
UNKNOWN_ROW_COLOR = "FDECEA"      # Light red for unknown/security rows


def _col_width(values: list, header: str, padding: int = 4) -> float:
    """Estimate column width from content."""
    max_len = max((len(str(v)) for v in values), default=0)
    return max(len(header), max_len) + padding


def export_to_excel(log_path: str = LOG_FILE_PATH) -> str | None:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    """
    Export the attendance CSV to a formatted Excel file.

    Returns the output filename on success, or ``None`` on failure.
    """
    if not os.path.exists(log_path):
        print(f"[ERROR] Log file not found: '{log_path}'")
        print("        Run run_recognition.py first to generate attendance data.")
        return None

    df = pd.read_csv(log_path)
    if df.empty:
        print("[INFO] Log file is empty — nothing to export.")
        try:
            os.remove(log_path)
        except OSError:
            pass
        return None

    # Ensure Photo column exists (may be absent in old logs)
    if "Snapshot" not in df.columns:
        df["Snapshot"] = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Log"

    # --- Header styling ---
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor=HEADER_COLOR)
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Columns: Name, Date, Time, Confidence, Photo
    display_cols = ["Name", "Date", "Time", "Confidence"]
    headers      = display_cols + ["Photo"]

    for col_idx, header in enumerate(headers, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center_align
        cell.border     = thin_border

    ws.row_dimensions[1].height = 22

    # --- Data rows ---
    photo_col_idx = len(headers)   # Last column = Photo

    for row_idx, (_, record) in enumerate(df.iterrows(), start=2):
        alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR) if row_idx % 2 == 0 else None

        is_unknown = str(record.get("Name", "")).strip() == "Unknown"
        row_fill   = PatternFill("solid", fgColor=UNKNOWN_ROW_COLOR) if is_unknown else alt_fill

        for col_idx, col_name in enumerate(display_cols, start=1):
            val  = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.alignment = center_align
            cell.border    = thin_border
            if is_unknown:
                cell.fill = row_fill
                cell.font = Font(bold=True, color="C0392B")  # red text for unknown
            elif alt_fill:
                cell.fill = alt_fill

        # Photo cell (placeholder text; image inserted below)
        photo_cell            = ws.cell(row=row_idx, column=photo_col_idx, value="")
        photo_cell.alignment  = center_align
        photo_cell.border     = thin_border
        if row_fill:
            photo_cell.fill = row_fill

        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PX

        # --- Embed snapshot ---
        snapshot_path = str(record.get("Snapshot", "")).strip()
        if snapshot_path and os.path.exists(snapshot_path):
            try:
                from PIL import Image as PILImage
                pil_img = PILImage.open(snapshot_path)
                pil_img.thumbnail((THUMB_W, THUMB_H))
                thumb_path = snapshot_path + "_thumb.png"
                pil_img.save(thumb_path, format="PNG")

                xl_img = XLImage(thumb_path)
                # Anchor: column letter + row number
                col_letter = get_column_letter(photo_col_idx)
                ws.add_image(xl_img, f"{col_letter}{row_idx}")
            except Exception as exc:
                print(f"[WARNING] Could not embed image for row {row_idx}: {exc}")

    # --- Column widths ---
    col_data = {h: df[h].astype(str).tolist() if h in df.columns else []
                for h in display_cols}
    for col_idx, col_name in enumerate(display_cols, start=1):
        width = _col_width(col_data[col_name], col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fixed width for photo column
    ws.column_dimensions[get_column_letter(photo_col_idx)].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Save ---
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = os.path.join(OUTPUT_DIR, f"Attendance_Export_{timestamp}.xlsx")
    wb.save(output_file)

    # Clean up temp log after successful export
    try:
        os.remove(log_path)
        print(f"[INFO] Temporary log '{log_path}' removed.")
    except OSError:
        pass

    # Clean up: delete thumbnails AND original snapshots after embedding
    deleted = 0
    for _, record in df.iterrows():
        snap_path = str(record.get("Snapshot", "")).strip()
        thumb_path = snap_path + "_thumb.png"
        for path in (thumb_path, snap_path):
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                    deleted += 1
                except OSError:
                    pass

    if deleted:
        print(f"[INFO] Deleted {deleted} snapshot file(s) from snapshots/ folder.")

    print(f"[INFO] Excel export complete -> '{output_file}'")
    return output_file


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    result = export_to_excel()
    if result is None:
        sys.exit(1)
